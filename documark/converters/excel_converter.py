"""
Excel 转换器
"""

import csv
from pathlib import Path
from typing import Dict, Any, List

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .base import BaseConverter


class ExcelConverter(BaseConverter):
    """
    Excel表格转换器
    
    将.xls, .xlsx, .csv文件转换为Markdown格式
    """
    
    CATEGORY = "表格"
    
    def convert(self, file_path: Path) -> str:
        """
        将Excel/CSV转换为Markdown
        
        Args:
            file_path: Excel或CSV文件路径
            
        Returns:
            Markdown格式的文本内容
        """
        ext = file_path.suffix.lower()
        
        if ext == '.csv':
            return self._convert_csv(file_path)
        elif ext in ['.xls', '.xlsx']:
            return self._convert_excel(file_path)
        else:
            raise ValueError(f"不支持的文件格式: {ext}")
    
    def _convert_csv(self, file_path: Path) -> str:
        """转换CSV文件"""
        content_parts = []
        
        # 添加文档标题
        content_parts.append(self._create_header(f"表格: {file_path.stem}", 1))
        content_parts.append("---\n\n")
        
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            rows = list(reader)
        
        if rows:
            headers = rows[0]
            data_rows = rows[1:] if len(rows) > 1 else []
            content_parts.append(self._create_table(headers, data_rows))
        
        return "".join(content_parts)
    
    def _convert_excel(self, file_path: Path) -> str:
        """转换Excel文件"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("请安装 openpyxl: pip install openpyxl")
        
        content_parts = []
        
        # 添加文档标题
        content_parts.append(self._create_header(f"工作簿: {file_path.stem}", 1))
        content_parts.append("---\n\n")
        
        workbook = load_workbook(file_path, data_only=True)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            content_parts.append(self._create_header(f"工作表: {sheet_name}", 2))
            
            # 提取数据
            rows_data = []
            for row in sheet.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                # 跳过空行
                if any(cell.strip() for cell in row_data):
                    rows_data.append(row_data)
            
            if rows_data:
                # 检测表头（第一行）
                headers = rows_data[0]
                data_rows = rows_data[1:] if len(rows_data) > 1 else []
                
                content_parts.append(self._create_table(headers, data_rows))
            
            content_parts.append("\n")
        
        return "".join(content_parts)
